import os
import sys
import pandas as pd
from PIL import Image, ImageDraw, ImageFont
import openpyxl
from openpyxl.drawing.image import Image as OpenpyxlImage
import tempfile
import subprocess

class ExcelToImageConverter:
    def __init__(self, excel_file_path):
        self.excel_file_path = excel_file_path
        
    def method_1_xlwings(self, sheet_name=None, output_path=None):
        """Método 1: Usando xlwings (Windows com Excel instalado)"""
        try:
            import xlwings as xw
            
            # Abre o Excel
            app = xw.App(visible=False)
            wb = app.books.open(self.excel_file_path)
            
            if sheet_name:
                ws = wb.sheets[sheet_name]
            else:
                ws = wb.sheets[0]
            
            # Encontra a área usada
            used_range = ws.used_range
            
            if not output_path:
                output_path = f"{sheet_name or 'planilha'}.png"
            
            # Copia como imagem
            used_range.api.CopyPicture(Format=2)  # xlBitmap
            
            # Salva a imagem usando PIL
            from PIL import ImageGrab
            img = ImageGrab.grabclipboard()
            if img:
                img.save(output_path, 'PNG', dpi=(300, 300))
                print(f"SUCESSO Método xlwings: Imagem salva como {output_path}")
                
            # Fecha o Excel
            wb.close()
            app.quit()
            
            return output_path
            
        except ImportError:
            print("ERRO xlwings não instalado")
            return None
        except Exception as e:
            print(f"ERRO no método xlwings: {e}")
            return None
    
    def method_2_excel_com(self, sheet_name=None, output_path=None):
        """Método 2: Usando COM automation (Windows)"""
        try:
            import win32com.client as win32
            from PIL import ImageGrab
            import time
            
            # Inicia Excel
            excel = win32.gencache.EnsureDispatch('Excel.Application')
            excel.Visible = False
            excel.DisplayAlerts = False
            
            # Abre workbook
            wb = excel.Workbooks.Open(os.path.abspath(self.excel_file_path))
            
            if sheet_name:
                ws = wb.Worksheets(sheet_name)
            else:
                ws = wb.Worksheets(1)
            
            ws.Activate()
            
            # Seleciona área usada
            used_range = ws.UsedRange
            used_range.Select()
            
            # Copia como imagem
            used_range.CopyPicture(Format=2)  # xlBitmap
            
            if not output_path:
                output_path = f"{sheet_name or 'planilha'}.png"
            
            # Captura da área de transferência
            time.sleep(1)  # Aguarda um pouco
            img = ImageGrab.grabclipboard()
            if img:
                img.save(output_path, 'PNG', dpi=(300, 300))
                print(f"SUCESSO Método COM: Imagem salva como {output_path}")
            
            # Fecha Excel
            wb.Close(SaveChanges=False)
            excel.Quit()
            
            return output_path
            
        except ImportError:
            print("ERRO pywin32 não instalado")
            return None
        except Exception as e:
            print(f"ERRO no método COM: {e}")
            return None
    
    def method_3_libreoffice(self, sheet_name=None, output_path=None):
        """Método 3: Usando LibreOffice via linha de comando"""
        try:
            if not output_path:
                output_path = f"{sheet_name or 'planilha'}.png"
            
            # Converte para PDF primeiro, depois para imagem
            temp_dir = tempfile.mkdtemp()
            pdf_path = os.path.join(temp_dir, "temp.pdf")
            
            # Comando LibreOffice
            cmd = [
                "libreoffice",
                "--headless",
                "--convert-to", "pdf",
                "--outdir", temp_dir,
                self.excel_file_path
            ]
            
            result = subprocess.run(cmd, capture_output=True, text=True)
            
            if result.returncode == 0:
                # Converte PDF para imagem usando PIL
                try:
                    from pdf2image import convert_from_path
                    images = convert_from_path(pdf_path, dpi=300)
                    if images:
                        images[0].save(output_path, 'PNG')
                        print(f"SUCESSO Método LibreOffice: Imagem salva como {output_path}")
                        return output_path
                except ImportError:
                    print("ERRO pdf2image não instalado")
            
            return None
            
        except Exception as e:
            print(f"ERRO no método LibreOffice: {e}")
            return None
    
    def method_4_aspose(self, sheet_name=None, output_path=None):
        """Método 4: Usando Aspose.Cells (biblioteca comercial)"""
        try:
            from aspose.cells import Workbook, ImageType, ImageOrPrintOptions
            
            # Carrega workbook
            workbook = Workbook(self.excel_file_path)
            
            if sheet_name:
                worksheet = workbook.getWorksheets().get(sheet_name)
            else:
                worksheet = workbook.getWorksheets().get(0)
            
            if not output_path:
                output_path = f"{sheet_name or 'planilha'}.png"
            
            # Configurações de imagem
            options = ImageOrPrintOptions()
            options.setImageType(ImageType.PNG)
            options.setQuality(100)
            options.setResolution(300)
            
            # Converte para imagem
            worksheet.toImage(output_path, options)
            print(f"SUCESSO Método Aspose: Imagem salva como {output_path}")
            
            return output_path
            
        except ImportError:
            print("ERRO aspose-cells não instalado")
            return None
        except Exception as e:
            print(f"ERRO no método Aspose: {e}")
            return None
    
    def method_5_improved_openpyxl(self, sheet_name=None, output_path=None):
        """Método 5: OpenPyXL melhorado com renderização mais precisa"""
        try:
            from openpyxl import load_workbook
            from openpyxl.utils import get_column_letter
            from PIL import Image, ImageDraw, ImageFont
            import math
            
            wb = load_workbook(self.excel_file_path, data_only=False)
            
            if sheet_name:
                ws = wb[sheet_name]
            else:
                ws = wb.active
            
            if not output_path:
                output_path = f"{sheet_name or 'planilha'}.png"
            
            # Calcula dimensões reais baseado no Excel
            EXCEL_POINT_TO_PIXEL = 1.33  # Conversão mais precisa
            DEFAULT_COL_WIDTH = 64  # Largura padrão em pixels
            DEFAULT_ROW_HEIGHT = 20  # Altura padrão em pixels
            
            # Dimensões das colunas
            col_widths = {}
            for col in range(1, ws.max_column + 1):
                col_letter = get_column_letter(col)
                col_dim = ws.column_dimensions[col_letter]
                if col_dim.width:
                    # Conversão mais precisa da largura
                    col_widths[col] = int(col_dim.width * 7.5)  # Fator de conversão do Excel
                else:
                    col_widths[col] = DEFAULT_COL_WIDTH
            
            # Dimensões das linhas
            row_heights = {}
            for row in range(1, ws.max_row + 1):
                row_dim = ws.row_dimensions[row]
                if row_dim.height:
                    row_heights[row] = int(row_dim.height * EXCEL_POINT_TO_PIXEL)
                else:
                    row_heights[row] = DEFAULT_ROW_HEIGHT
            
            # Calcula tamanho total
            total_width = sum(col_widths.values())
            total_height = sum(row_heights.values())
            
            # Cria imagem com fundo branco
            img = Image.new('RGB', (total_width, total_height), (255, 255, 255))
            draw = ImageDraw.Draw(img)
            
            # Desenha grade de fundo (opcional)
            current_y = 0
            for row in range(1, ws.max_row + 1):
                current_x = 0
                for col in range(1, ws.max_column + 1):
                    x1, y1 = current_x, current_y
                    x2, y2 = current_x + col_widths[col], current_y + row_heights[row]
                    
                    # Desenha borda de célula muito sutil
                    draw.rectangle([x1, y1, x2, y2], outline=(240, 240, 240), width=1)
                    
                    current_x += col_widths[col]
                current_y += row_heights[row]
            
            # Desenha conteúdo das células
            current_y = 0
            for row in range(1, ws.max_row + 1):
                current_x = 0
                
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row, column=col)
                    
                    x1, y1 = current_x, current_y
                    x2, y2 = current_x + col_widths[col], current_y + row_heights[row]
                    
                    # Cor de fundo da célula
                    if cell.fill and hasattr(cell.fill, 'start_color') and cell.fill.start_color.rgb:
                        bg_color = self._hex_to_rgb(cell.fill.start_color.rgb)
                        draw.rectangle([x1, y1, x2, y2], fill=bg_color)
                    
                    # Bordas da célula
                    if cell.border:
                        self._draw_precise_border(draw, x1, y1, x2, y2, cell.border)
                    
                    # Texto da célula
                    if cell.value is not None:
                        self._draw_cell_text(draw, cell, x1, y1, x2, y2)
                    
                    current_x += col_widths[col]
                
                current_y += row_heights[row]
            
            # Salva imagem em alta resolução
            img.save(output_path, 'PNG', dpi=(300, 300), optimize=True)
            print(f"SUCESSO Método OpenPyXL melhorado: Imagem salva como {output_path}")
            
            return output_path
            
        except Exception as e:
            print(f"ERRO no método OpenPyXL melhorado: {e}")
            return None
    
    def _hex_to_rgb(self, hex_color):
        """Converte cor hex para RGB"""
        if isinstance(hex_color, str) and len(hex_color) >= 6:
            hex_color = hex_color[-6:]  # Pega os últimos 6 caracteres
            try:
                return tuple(int(hex_color[i:i+2], 16) for i in (0, 2, 4))
            except:
                pass
        return (255, 255, 255)
    
    def _draw_precise_border(self, draw, x1, y1, x2, y2, border):
        """Desenha bordas com mais precisão"""
        border_color = (0, 0, 0)  # Preto
        
        if border.left and border.left.style:
            width = 2 if border.left.style == 'thick' else 1
            draw.line([(x1, y1), (x1, y2)], fill=border_color, width=width)
        
        if border.right and border.right.style:
            width = 2 if border.right.style == 'thick' else 1
            draw.line([(x2, y1), (x2, y2)], fill=border_color, width=width)
        
        if border.top and border.top.style:
            width = 2 if border.top.style == 'thick' else 1
            draw.line([(x1, y1), (x2, y1)], fill=border_color, width=width)
        
        if border.bottom and border.bottom.style:
            width = 2 if border.bottom.style == 'thick' else 1
            draw.line([(x1, y2), (x2, y2)], fill=border_color, width=width)
    
    def _draw_cell_text(self, draw, cell, x1, y1, x2, y2):
        """Desenha texto da célula com formatação precisa"""
        text = str(cell.value)
        
        # Fonte
        try:
            font_size = int(cell.font.size) if cell.font.size else 11
            font = self._get_system_font(font_size, cell.font.bold if cell.font else False)
        except:
            font = ImageFont.load_default()
        
        # Cor do texto
        if cell.font and cell.font.color and hasattr(cell.font.color, 'rgb'):
            text_color = self._hex_to_rgb(cell.font.color.rgb)
        else:
            text_color = (0, 0, 0)
        
        # Calcula posição do texto
        try:
            bbox = draw.textbbox((0, 0), text, font=font)
            text_width = bbox[2] - bbox[0]
            text_height = bbox[3] - bbox[1]
        except:
            text_width = len(text) * 8
            text_height = 14
        
        # Alinhamento horizontal
        cell_width = x2 - x1
        if cell.alignment and cell.alignment.horizontal == 'center':
            text_x = x1 + (cell_width - text_width) // 2
        elif cell.alignment and cell.alignment.horizontal == 'right':
            text_x = x2 - text_width - 3
        else:
            text_x = x1 + 3
        
        # Alinhamento vertical
        cell_height = y2 - y1
        if cell.alignment and cell.alignment.vertical == 'center':
            text_y = y1 + (cell_height - text_height) // 2
        elif cell.alignment and cell.alignment.vertical == 'bottom':
            text_y = y2 - text_height - 2
        else:
            text_y = y1 + 2
        
        # Desenha o texto
        draw.text((text_x, text_y), text, fill=text_color, font=font)
    
    def _get_system_font(self, size, bold=False):
        """Obtém fonte do sistema"""
        try:
            if os.name == 'nt':  # Windows
                font_name = "arial.ttf"
                font_path = f"C:/Windows/Fonts/{font_name}"
            else:  # Linux/Mac
                font_path = "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"
            
            if os.path.exists(font_path):
                return ImageFont.truetype(font_path, size)
        except:
            pass
        
        return ImageFont.load_default()
    
    def convert_to_image(self, sheet_name=None, output_path=None):
        """Tenta diferentes métodos em ordem de precisão"""
        print("Tentando conversao com layout exato...")
        
        methods = [
            ("xlwings (Excel + Windows)", self.method_1_xlwings),
            ("COM Automation (Windows)", self.method_2_excel_com),
            ("Aspose.Cells (Comercial)", self.method_4_aspose),
            ("LibreOffice", self.method_3_libreoffice),
            ("OpenPyXL Melhorado", self.method_5_improved_openpyxl)
        ]
        
        for method_name, method_func in methods:
            print(f"TENTANDO método: {method_name}")
            result = method_func(sheet_name, output_path)
            if result:
                print(f"SUCESSO com método: {method_name}")
                return result
        
        print("ERRO Nenhum método funcionou")
        return None

# Função simplificada para uso
def xlsx_to_image_exact(excel_file_path, output_path=None, sheet_name=None):
    """
    Converte arquivo XLSX para imagem mantendo layout EXATO
    
    Args:
        excel_file_path: Caminho para o arquivo Excel
        output_path: Caminho de saída da imagem
        sheet_name: Nome da planilha específica
    
    Returns:
        Caminho do arquivo de imagem criado
    """
    converter = ExcelToImageConverter(excel_file_path)
    return converter.convert_to_image(sheet_name, output_path)


# Execução via linha de comando
if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Uso: python excel_to_image.py <arquivo_excel> <saida_imagem>")
        sys.exit(1)
    
    excel_path = sys.argv[1]
    output_path = sys.argv[2]
    
    result = xlsx_to_image_exact(excel_path, output_path)
    
    if result:
        print(f"SUCCESS:{result}")
    else:
        print("ERROR:Conversão falhou")
        sys.exit(1)